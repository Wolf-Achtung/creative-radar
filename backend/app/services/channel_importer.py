from io import BytesIO
from urllib.parse import urlparse

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.models.entities import Channel, Market, Priority


def _cell(value) -> str:
    return str(value or '').strip()


def extract_handle(url: str) -> str:
    clean = _cell(url).rstrip('/')
    if not clean:
        return ''
    if not clean.startswith('http'):
        clean = 'https://' + clean
    path = urlparse(clean).path
    parts = [part for part in path.split('/') if part]
    return parts[0] if parts else ''


def guess_market(name: str, handle: str) -> Market:
    text = f'{name} {handle}'.lower()
    de_tokens = ['deutschland', 'germany', 'austria', 'dach', 'filmverleih', 'verleih', 'weltkino', 'tobis', 'leonine', 'constantin', 'studiocanalde', 'skydeutschland', 'wowtv']
    if handle.endswith('de') or handle.endswith('_de') or any(token in text for token in de_tokens):
        return Market.DE
    us_tokens = ['a24', 'warner', 'universal', 'sony', 'paramount', 'disney', 'marvel', 'pixar', 'hbo', 'hulu', 'netflix', 'primevideo', 'amazonmgm', '20th']
    if any(token in text for token in us_tokens):
        return Market.US
    return Market.INT


def guess_channel_type(name: str, handle: str) -> str:
    text = f'{name} {handle}'.lower()
    if any(token in text for token in ['netflix', 'prime', 'disneyplus', 'hbo', 'hulu', 'apple', 'sky', 'wow', 'discovery', 'paramountplus']):
        return 'Streamer'
    if any(token in text for token in ['plaion', 'game', 'games']):
        return 'Game Publisher'
    if any(token in text for token in ['studio', 'studios', 'pictures', 'warner', 'universal', 'sony', 'paramount', 'disney', 'marvel', 'pixar', '20th']):
        return 'Studio/Verleih'
    return 'Verleih/Produktion'


def guess_priority(name: str, handle: str, market: Market) -> Priority:
    text = f'{name} {handle}'.lower()
    high = ['a24', 'warner', 'universal', 'sony', 'paramount', 'disney', 'marvel', 'pixar', 'hbo', 'netflix', 'primevideo', 'apple', '20th', 'constantin', 'leonine', 'studiocanal', 'sky', 'wow']
    if any(token in text for token in high):
        return Priority.A
    return Priority.B if market == Market.DE else Priority.C


def import_channels_from_excel(session: Session, data: bytes) -> dict:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {'created': 0, 'updated': 0, 'skipped': 0}

    headers = [_cell(value).lower() for value in rows[0]]
    try:
        name_idx = headers.index('filmverleih')
    except ValueError:
        name_idx = 0
    try:
        url_idx = headers.index('instagram')
    except ValueError:
        url_idx = 1

    created = updated = skipped = 0
    for row in rows[1:]:
        name = _cell(row[name_idx] if len(row) > name_idx else '')
        url = _cell(row[url_idx] if len(row) > url_idx else '').rstrip('/')
        if not name or not url:
            skipped += 1
            continue
        handle = extract_handle(url)
        if not handle:
            skipped += 1
            continue
        if not url.startswith('http'):
            url = 'https://' + url
        market = guess_market(name, handle)
        channel_type = guess_channel_type(name, handle)
        priority = guess_priority(name, handle, market)
        channel = session.exec(select(Channel).where(Channel.handle == handle)).first()
        if channel:
            channel.name = name
            channel.url = url
            channel.market = market
            channel.channel_type = channel_type
            channel.priority = priority
            channel.active = True
            channel.mvp = True
            updated += 1
        else:
            channel = Channel(
                name=name,
                url=url,
                handle=handle,
                market=market,
                channel_type=channel_type,
                priority=priority,
                active=True,
                mvp=True,
            )
            created += 1
        session.add(channel)
    session.commit()
    return {'created': created, 'updated': updated, 'skipped': skipped}
